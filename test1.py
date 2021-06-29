import keyring
import openpyxl
import time
import logging
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By

Config={}
TransactionNumber=1
TransactionItem=[]
List=[]
RetryNumber=0



logging.basicConfig(filename="C://Users/Vivek Kumar/Documents/Sample.log",
                        format='%(asctime)s: %(levelname)s: %(message)s',
                        datefmt='%m/%d/%y %I:%M:%S %p'
                        )
logger = logging.getLogger()
logger.setLevel(logging.DEBUG)

class BusinessRuleException(Exception):
    def __init__(self, value):
        self.value = value
    def __str__(self):
        return (repr(self.value))



def init():
    try:
        SystemError="Nothing"
        if bool(Config):
            InitAllSettings()
            KillAllProcess()
            logger.info("******************Process Started***********************")
        InitAllApplication()
    except BusinessRuleException as error:
        SystemError=error.value
    except Exception as error:
        SystemError=error.value
    if SystemError is "Nothing":
        GetTransactionData()
    else:
        logger.info("System Error at Initialization: "+SystemError)
        EndProcess()

def KillAllProcess():
    pass
def InitAllApplication():
    logger.debug("Starting All Application")
    logger.debug("Logging into Acme.")
    isLoggedIn=AcmeLogin()
    if isLoggedIn=="Successful":
        logger.info("Login Successful")
    else:
        raise BusinessRuleException("Unable to Login")
def GetTransactionData():
    if(TransactionNumber<List.count):
        TransactionItem=List(TransactionNumber)
    else:
        TransactionItem=[]
def InitAllSettings():
    path = "Data/Config.xlsx"
    #path ="C:/Users/Vivek Kumar/PycharmProjects/REFramework/Data/Config.xlsx"
    workbook = openpyxl.load_workbook(path)
    for sheets in ("Settings","Constants","Assets"):
        sheet = workbook.get_sheet_by_name(sheets)
        rows = sheet.max_row
        for r in range(1, rows + 1):
            key=sheet.cell(row=r, column=1).value
            value=sheet.cell(row=r, column=2).value
            Config.update({key: value})


def SetTransactionStatus(in_Config,in_SystemError,in_BusinessRuleException,in_TransactionItem,io_RetryNumber,io_TransactionNumber):
    if(in_BusinessRuleException is "Nothing" and in_SystemError is "Nothing"):
        io_TransactionNumber=io_TransactionNumber+1
        io_RetryNumber=0
    else:
        if(in_BusinessRuleException isnot Nothing):
            io_TransactionNumber = io_TransactionNumber + 1
            io_RetryNumber = 0
        else:
            if(cint(in_Config("MaxRetryNumber"))>0):
                if(io_RetryNumber>=cint(in_Config("MaxRetryNumber"))):
                    io_TransactionNumber = io_TransactionNumber + 1
                    io_RetryNumber = 0
                else:
                    io_RetryNumber = io_RetryNumber+1
            else:
                io_TransactionNumber = io_TransactionNumber + 1

import keyring
#keyring.set_password("system", "username", "password")
print(keyring.get_password("system", "username"))
pip install keyring


import Framework.InitAllSettings
import Framework.KillAllProcesses
import Framework.InitAllApplication
import Framework.GetTransactionData
import Framework.SetTransactionStatus
import Framework.CloseAllApplications
import Process
import Logs
import logging


logging.basicConfig(filename="C://Users/Vivek Kumar/Documents/Sample.log",
                    format='%(asctime)s: %(levelname)s: %(message)s',
                    datefmt='%m/%d/%y %I:%M:%S %p',
                    level=logging.DEBUG
                    )
logger=logging.getLogger()
logger.setLevel(logging.DEBUG)

Config={}
shouldStop=False
TransactionItem=[]
TransationNumber=1

def init():
    try:
        SystemError="Nothing"
        if bool(Config):
            Framework.InitAllSettings.InitAllSettings("Data\Config.xlsx",{"Settings", "Constants","Assets"},Config)
            Framework.KillAllProcesses.KillAllProcesses()
            logger.info("******************Process Started***********************")
        Framework.InitAllApplication.InitAllApplication()
    except:
        SystemError="Exception"
    if SystemError is "Nothing":
        GetTransactionData()
    else:
        EndProcess()

def GetTransactionData():
        Framework.GetTransactionData.GetTransactionData()
    if TransactionItem ==[]:
        logger.info("Process finished due to no more transaction data.")
        EndProcess()
    else:
        logger.info("Processing Transaction Number: ",TransationNumber)
        ProcessTransaction()

def ProcessTransaction():
    try:
        BusinessRuleException="Nothing"
        Process.Process()
    except
        BusinessRuleException=exception
    finally:
        try:
            Framework.SetTransactionStatus.SetTransactionStatus()
        except:
            logger.info("Set Transaction Status Failed.")
    if SystemError=="NotNothing":
        init()
    elif BusinessRuleException=="NotNothing":
        GetTransactionData()
    else:
        GetTransactionData()
def EndProcess():
    try:
        Framework.CloseAllApplications.CloseAllApplications()
    except:
        logger.warn("Unable to Close the process Normally.")
        Framework.KillAllProcesses.KillAllProcesses()
init()

def Logs(level,message):

def AcmeLogin():
    pass