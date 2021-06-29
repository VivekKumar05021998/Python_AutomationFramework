import keyring
import openpyxl
import time
import logging
import subprocess
import datetime
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By





Config={}
TransactionNumber=0
TransactionItem=[]
List=[]
RetryNumber=0
SystemError="Nothing"
todayDate=datetime.date.today()


logging.basicConfig(filename="Logs/Log--"+str(todayDate)+".log",
                    format='%(asctime)s: %(levelname)s: %(message)s',
                    datefmt='%m/%d/%y %I:%M:%S %p'
                    )
logger = logging.getLogger()
logger.setLevel(logging.INFO)

class BusinessRuleException(Exception):
    def __init__(self, value):
        self.value = value
    def __str__(self):
        return (repr(self.value))



logger.info("******************Process Started***********************")

def init(Config,TransactionNumber,TransactionItem,List,RetryNumber,SystemError,driver):
    try:
        SystemError="Nothing"
        if not bool(Config):
            Config=InitAllSettings(Config)
            logger.info("Reading Data from Config Successful.")
            path = "Process_Case 2.xlsx"
            workbook = openpyxl.load_workbook(path)
            sheet = workbook.get_sheet_by_name("Sheet1")
            rows = sheet.max_row
            for r in range(1, rows + 1):
                name = sheet.cell(row=r, column=2).value
                url = sheet.cell(row=r, column=3).value
                List.append([name,url])
            print(List)
        driver=InitAllApplication(Config)
    except BusinessRuleException as error:
        SystemError=error.value
        print(SystemError)
    except Exception:
        SystemError="Error"
    if SystemError is "Nothing":
        print("GetTransactionData")
        Config,TransactionNumber,TransactionItem,List,RetryNumber,SystemError,driver=GetTransactionData(Config,TransactionNumber,TransactionItem,List,RetryNumber,SystemError,driver)
    else:
        logger.info("System Error at Initialization: "+SystemError)
        print("EndProcess")
        EndProcess(Config,driver)

def GetTransactionData(Config,TransactionNumber,TransactionItem,List,RetryNumber,SystemError,driver):
    if(TransactionNumber<len(List)):
        TransactionItem=List[TransactionNumber]
        print(TransactionItem)
    else:
        TransactionItem=[]
    if TransactionItem ==[]:
        logger.info("Process finished due to no more transaction data.")
        EndProcess(Config,driver)
    else:
        logger.info("Processing Transaction Number: "+str(TransactionNumber))
        Config,TransactionNumber,TransactionItem,List,RetryNumber,SystemError,driver=ProcessTransaction(Config,TransactionNumber,TransactionItem,List,RetryNumber,SystemError,driver)
    return Config,TransactionNumber,TransactionItem,List,RetryNumber,SystemError,driver

def ProcessTransaction(Config,TransactionNumber,TransactionItem,List,RetryNumber,SystemError,driver):
    try:
        SystemError="Nothing"
        print("Inside Process Transaction.")
        Process(Config,TransactionItem,driver)
    except BusinessRuleException as error:
        SystemError=error.value
        print("Inside BusinessRuleException")
        print(SystemError)
    except Exception as error:
        SystemError="Error"
        print("Inside Exception.")
    finally:
        try:
            TransactionNumber,RetryNumber=SetTransactionStatus(TransactionNumber,RetryNumber,SystemError)
        except:
            logger.info("Set Transaction Status Failed.")
    if SystemError=="Nothing":
        print("Gotig to Get transaction data.")
        GetTransactionData(Config, TransactionNumber, TransactionItem, List, RetryNumber, SystemError,driver)
    elif SystemError=="Error":
        print("Gotig to Init data.")
        init(Config, TransactionNumber, TransactionItem, List, RetryNumber, SystemError, driver)
    else:
        print("Going into Get Transaction State.")
        GetTransactionData(Config, TransactionNumber, TransactionItem, List, RetryNumber, SystemError, driver)
    return Config,TransactionNumber,TransactionItem,List,RetryNumber,SystemError,driver

def EndProcess(Config,driver):
    try:
        CloseAllApplication(Config,driver)
    except:
        logger.warn("Unable to Close the process Normally.")
        KillAllProcess()





def InitAllSettings(Config):
    path = "Config.xlsx"
    #path ="C:/Users/Vivek Kumar/PycharmProjects/REFramework/Data/Config.xlsx"
    workbook = openpyxl.load_workbook(path)
    for sheets in ("Settings","Constants","Assets"):
        sheet = workbook.get_sheet_by_name(sheets)
        rows = sheet.max_row
        for r in range(1, rows + 1):
            key=sheet.cell(row=r, column=1).value
            value=sheet.cell(row=r, column=2).value
            if key==None:
                pass
            else:
                Config.update({key: value})
    return Config

def KillAllProcess():
    subprocess.call([r'TaskKill.bat'])
    time.sleep(5)

def InitAllApplication(Config):
    logger.debug("Starting All Application")
    logger.debug("Logging into Screener.")
    isLoggedIn,driver=ScreenerLogin(Config)
    if isLoggedIn=="Successful":
        logger.info("Login Successful")
        return driver
    else:
        raise BusinessRuleException("Unable to Login")
        return 0


def GetAppCredential(Config):
    password=keyring.get_password(Config["ScreenerCredentialName"],Config["Username"])
    return password

def SetTransactionStatus(TransactionNumber, RetryNumber,SystemError):
    if (SystemError is "Nothing"):
        TransactionNumber = TransactionNumber + 1
        RetryNumber = 0
    elif (SystemError is "Error"):
        if (int(Config["MaxRetryNumber"]) > 0):
            if (int(RetryNumber) >= int(Config["MaxRetryNumber"])):
                TransactionNumber = TransactionNumber + 1
                RetryNumber = 0
            else:
                RetryNumber = RetryNumber + 1
        else:
            TransactionNumber = TransactionNumber + 1
    else:
        TransactionNumber = TransactionNumber + 1
        RetryNumber = 0
    return TransactionNumber,RetryNumber

def CloseAllApplication(Config,driver):
    logger.debug("Closing the Application")
    ScreenerLogout(Config,driver)



def Process(Config,TransactionItem,driver):
    print("Inside Process.")
    print(TransactionItem)
    if TransactionItem[1]=="NOVIGO":
        raise BusinessRuleException("Novigo Solutions.")
    driver.find_element_by_xpath("//*[@aria-label='Search Company'][@type='search']").clear()
    driver.find_element_by_xpath("//*[@aria-label='Search Company'][@type='search']").send_keys(TransactionItem[1])
    time.sleep(5)
    print("Leaving Process")
def ScreenerLogin(Config):
    print("Before retry Number")
    retryNumber = int(Config["MaxRetryNumber"]+1)
    print(retryNumber)
    print("After Retry Numberr")
    for ik in range(retryNumber):
        print("After For loop")
        try:
            print("Before KillAllProcess")
            KillAllProcess()
            print("After Kill All APrdm")
            time.sleep(8)
            chrome_options = webdriver.ChromeOptions()
            chrome_options.add_argument("--incognito")
            driver = webdriver.Chrome(executable_path="C:\chromedriver.exe", chrome_options=chrome_options)
            driver.get("https://www.screener.in/")
            driver.maximize_window()
            driver.implicitly_wait(10)
            driver.find_element_by_xpath("//*[contains(text(),'Login')]").click()
            driver.find_element_by_xpath("//*[@type='text'][@name='username']").send_keys(Config["Username"])
            driver.find_element_by_xpath("//*[@type='password'][@name='password']").send_keys(GetAppCredential(Config))
            print("After Password.")
            driver.find_element_by_xpath("//*[@type='submit'][@class='button-primary']").click()
            print("After Login Button")
            logout = driver.find_element_by_xpath("//*[@aria-label='Logout'][@type='submit']").is_displayed()
            if not logout:
                raise Exception
                print("Throwing Error.")
        except:
            print("Inside Except"+str(ik))
        else:
            break
    if (logout):
        print("SuccessFul")
        return "Successful",driver
    else:
        return "Failed."
def ScreenerLogout(Config,driver):
    driver.find_element_by_xpath("//*[@aria-label='Logout'][@type='submit']").click()
    driver.close()


init(Config,TransactionNumber,TransactionItem,List,RetryNumber,SystemError,"driver")